from lib.db.Custom_SQLite import Custom_SQLite as cS
from lib.db.ConnectPostgree import ConnectDb as pg

import sqlite3
import csv
import xlsxwriter
import pandas as pd
import openpyxl

def init():
    config_db = cS(config_db_url)
    config_db.create_connection()

def sumIgnoreNone(number, reciver):
    if number is not None:
        reciver = reciver + number
    return reciver

def isANumber(myVariable):
    if type(myVariable) == int or type(myVariable) == float:
        return True
    else:
        return False

def calcAllDataForNils():

    config_db_url = r"db\config.sqlite"
    pg_url = "icra.loading.net"
    pg_user = "traca_user"
    pg_pass = "EdificiH2O!"
    pg_db = "traca_1"

    edar_compounds_csv = "inputs/EDAR_DBO_TN_PT_media2.csv"
    edar_population_csv = "inputs/dp_population_country.csv"
    edar_analitiques_xlsx = "inputs/edars_analitiques_sistemes_1.xlsx"
    edar_ptr_xlsx = "inputs/prtr_edars.xlsx"

    connection = pg(pg_url, pg_db, pg_user, pg_pass)
    wwtps = connection.getAllWWTP()
    industries = connection.getIndustriesToEdar()
    acceptedWWTP = []
    dicWWTP = {}

    # Fields of edar compounts
    # cabal_diari m3/dia
    # unidades mg/l
    listOfEDARCompounds = {}
    with open(edar_compounds_csv, encoding='utf8', newline='') as csvEDARCompounts:
        reader = csv.reader(csvEDARCompounts, delimiter=';')
        isFirst = True
        for row in reader:
            if isFirst:
                isFirst = False
            else:
                listOfEDARCompounds[row[0]] = {
                    "eu_code": row[0],
                    "dc_code": row[1],
                    "nom": row[2],
                    "industries": [],
                    "industriesTotalInfluent": {
                        "amoni": 0,
                        "dbo": 0,
                        "fosfats": 0,
                        "toc": 0,
                        "fosfor": 0,
                        "nitrats": 0,
                        "nitrogen_org": 0,
                        "cabal": 0,
                    },
                    "industriesTotalEfluent": {
                        "amoni": 0,
                        "dbo": 0,
                        "fosfats": 0,
                        "toc": 0,
                        "fosfor": 0,
                        "nitrats": 0,
                        "nitrogen_org": 0,
                        "cabal": 0,
                    },
                    "configuration": [],
                    "efluent":{
                        "dbo": None,
                        "nitrogen": None,
                        "fosfor": None,
                        "arsènic i compostos com as µg as/l": None,
                        "cadmi i compostos com cd µg cd/l": None,
                        "crom i compostos com cr µg cr/l": None,
                        "coure i compostos com cu µg cu/l": None,
                        "mercuri i compostos com hg µg hg/l": None,
                        "niquel i compostos com ni µg ni/l": None,
                        "plom i compostos com pb µg pb/l": None,
                        "zenc i compostos com zn µg zn/l": None,
                        "cianurs com cn total µg cn/l": None,
                        "fluorurs com f total µg f/l": None,
                        "clorurs com cl total µg cl/l": None,
                        "1,2 dicloroetà µg/l": None,
                        "diclormetà µg/l": None,
                        "tetracloretilè per µg/l": None,
                        "tetraclormetà tcm tetracloruro de carbono µg/l": None,
                        "tricloretilètri µg/l": None,
                        "triclormetà cloroformo µg/l": None,
                        "hexaclorbenzè µg/l": None,
                        "lindà gamma-hch µg/l": None,
                        "pentaclorfenol pcp µg/l": None,
                        "policlorbifenils pcbs µg/l": None,
                        "benzè µg/l": None,
                        "toluè µg/l": None,
                        "xilens µg/l": None,
                        "naftalè µg/l": None,
                        "fluorantè µg/l": None,
                        "benzo g, h, i, perilè µg/l": None,
                        "suma hap µg/l": None,
                        "atrazina µg/l": None,
                        "simazina µg/l": None,
                        "isoproturon µg/l": None,
                        "diuron µg/l": None,
                        "nonilfenol etoxilats np/np1,2eo µg/l": None,
                        "octilfenol etoxilatsop/op1,2eo µg/l": None,
                        "ftalat de bis 2-etilhexil deph µg/l": None,
                        "index de fenols µg fenol/l": None,
                        "compostos orgànics halogenats aox µgcl / l": None,
                        "tributil estany tbt µg tbt/l": None,
                        "tributil estany i els seus compostos mbt+dbt+tbt com tbt µg tb": None,
                        "trifenil estany tpht µg tpht/l": None,
                        "trifenil estany i els seus compostosmpht+dpht+tpht comtpht µg": None,
                        "compostos organoestannics totalscom sn µg sn/l": None,
                        "conductivitat 2Noneºc µs/cm": None,
                        "m+p-xilè µg/l": None,
                        "o-xilè µg/l": None,
                        "benzoapirè µg/l": None,
                        "benzobfluorantè µg/l": None,
                        "benzokfluorantè µg/l": None,
                        "indè1.2.3pirè µg/l": None,
                        "pcb 28 µg/l": None,
                        "pcb 52 µg/l": None,
                        "pcb 1None1 µg/l": None,
                        "pcb 118 µg/l": None,
                        "pcb 153 µg/l": None,
                        "pcb 138 µg/l": None,
                        "pcb 18None µg/l": None,
                        "ph u. ph": None,
                        "toc µg/l": None,
                        "dqo no decantada µg/l": None,
                        "nitrogen kjeldahl µg/l n": None,
                        "nitrogen total µg/l n": None,
                        "nitrats + nitrits µg/l n": None,
                        "4-nonilfenol dietoxilado": None,
                        "4-nonilfenol monoetoxilado": None,
                        "4-octilfenol dietoxilado": None,
                        "4-octilfenol monoetoxilado µg/l": None,
                        "compuestos organoestánnicos µg/l": None,
                        "dibutilestaño µg/l": None,
                        "difenilestaño µg/l": None,
                        "monobutilestaño µg/l": None,
                        "monofenilestaño µg/l": None,
                        "trifenilestaño µg/l": None,
                        "npeoe=None": None,
                        "t-octilfenol": None,
                        "dbo 5 dies µg o2/l": None,
                        "propazina µg/l": None,
                        "tertbutilazina µg/l": None,
                        "terbutrin µg/l": None,
                        "cibutrina µg/l": None,
                        "suma triazines µg/l": None,
                        "fòsfor total µg p/l": None
                    }
                }

    #

    with open(edar_population_csv, encoding='utf8', newline='') as csvEDARPopulation:
        reader = csv.reader(csvEDARPopulation, delimiter=',')
        isFirst = True
        for row in reader:
            if isFirst:
                isFirst = False
            else:
                if row[0] in listOfEDARCompounds:
                    listOfEDARCompounds[row[0]]["population"] = row[6]

    # ('FOMENTO AGRICOLA Y GANADERO, SL', 'ES9080010001010E', None, 750.0, None, None, 50.0, None, None, 651.0)



    for industry in industries:
        if industry[1] in listOfEDARCompounds:
            if industry[10] is not None and industry[10] > 0:
                
                actualIndustry = {
                    "nom": industry[0],
                    "tipus": industry[2],
                    "amoni": industry[3],
                    "dbo": industry[4],
                    "fosfats": industry[5],
                    "toc": industry[6],
                    "fosfor": industry[7],
                    "nitrats": industry[8],
                    "nitrogen_org": industry[9],
                    "cabal": industry[10],
                }

                # print(listOfEDARCompounds[industry[1]]["industries"])
                
                if actualIndustry["amoni"] is not None:
                    actualIndustry["amoni"] = industry[3]*(industry[10]*1000)
                if actualIndustry["dbo"] is not None:
                    actualIndustry["dbo"] = industry[4]*(industry[10]*1000)
                if actualIndustry["fosfats"] is not None:
                    actualIndustry["fosfats"] = industry[5]*(industry[10]*1000)
                if actualIndustry["toc"] is not None:
                    actualIndustry["toc"] = industry[6]*(industry[10]*1000)
                if actualIndustry["fosfor"] is not None:
                    actualIndustry["fosfor"] = industry[7]*(industry[10]*1000)
                if actualIndustry["nitrats"] is not None:
                    actualIndustry["nitrats"] = industry[8]*(industry[10]*1000)
                if actualIndustry["nitrogen_org"] is not None:
                    actualIndustry["nitrogen_org"] = industry[9]*(industry[10]*1000)
                if actualIndustry["cabal"] is not None:
                    actualIndustry["cabal"] = industry[10]*1000

                listOfEDARCompounds[industry[1]]["industries"].append(actualIndustry)
                
                if industry[2] in ["Abocament", "Depuradora", "Entrada EDAR"]:
                    listOfEDARCompounds[industry[1]]["industriesTotalInfluent"]["amoni"] = sumIgnoreNone(actualIndustry["amoni"], listOfEDARCompounds[industry[1]]["industriesTotalInfluent"]["amoni"])
                    listOfEDARCompounds[industry[1]]["industriesTotalInfluent"]["dbo"] = sumIgnoreNone(actualIndustry["dbo"], listOfEDARCompounds[industry[1]]["industriesTotalInfluent"]["dbo"])
                    listOfEDARCompounds[industry[1]]["industriesTotalInfluent"]["fosfats"] = sumIgnoreNone(actualIndustry["fosfats"], listOfEDARCompounds[industry[1]]["industriesTotalInfluent"]["fosfats"])
                    listOfEDARCompounds[industry[1]]["industriesTotalInfluent"]["toc"] = sumIgnoreNone(actualIndustry["toc"], listOfEDARCompounds[industry[1]]["industriesTotalInfluent"]["toc"])
                    listOfEDARCompounds[industry[1]]["industriesTotalInfluent"]["fosfor"] = sumIgnoreNone(actualIndustry["fosfor"], listOfEDARCompounds[industry[1]]["industriesTotalInfluent"]["fosfor"])
                    listOfEDARCompounds[industry[1]]["industriesTotalInfluent"]["nitrats"] = sumIgnoreNone(actualIndustry["nitrats"], listOfEDARCompounds[industry[1]]["industriesTotalInfluent"]["nitrats"])
                    listOfEDARCompounds[industry[1]]["industriesTotalInfluent"]["nitrogen_org"] = sumIgnoreNone(actualIndustry["nitrogen_org"], listOfEDARCompounds[industry[1]]["industriesTotalInfluent"]["nitrogen_org"])
                    listOfEDARCompounds[industry[1]]["industriesTotalInfluent"]["cabal"] = sumIgnoreNone(actualIndustry["cabal"], listOfEDARCompounds[industry[1]]["industriesTotalInfluent"]["cabal"])
                else:
                    listOfEDARCompounds[industry[1]]["industriesTotalEfluent"]["amoni"] = sumIgnoreNone(actualIndustry["amoni"], listOfEDARCompounds[industry[1]]["industriesTotalEfluent"]["amoni"])
                    listOfEDARCompounds[industry[1]]["industriesTotalEfluent"]["dbo"] = sumIgnoreNone(actualIndustry["dbo"], listOfEDARCompounds[industry[1]]["industriesTotalEfluent"]["dbo"])
                    listOfEDARCompounds[industry[1]]["industriesTotalEfluent"]["fosfats"] = sumIgnoreNone(actualIndustry["fosfats"], listOfEDARCompounds[industry[1]]["industriesTotalEfluent"]["fosfats"])
                    listOfEDARCompounds[industry[1]]["industriesTotalEfluent"]["toc"] = sumIgnoreNone(actualIndustry["toc"], listOfEDARCompounds[industry[1]]["industriesTotalEfluent"]["toc"])
                    listOfEDARCompounds[industry[1]]["industriesTotalEfluent"]["fosfor"] = sumIgnoreNone(actualIndustry["fosfor"], listOfEDARCompounds[industry[1]]["industriesTotalEfluent"]["fosfor"])
                    listOfEDARCompounds[industry[1]]["industriesTotalEfluent"]["nitrats"] = sumIgnoreNone(actualIndustry["nitrats"], listOfEDARCompounds[industry[1]]["industriesTotalEfluent"]["nitrats"])
                    listOfEDARCompounds[industry[1]]["industriesTotalEfluent"]["nitrogen_org"] = sumIgnoreNone(actualIndustry["nitrogen_org"], listOfEDARCompounds[industry[1]]["industriesTotalEfluent"]["nitrogen_org"])
                    listOfEDARCompounds[industry[1]]["industriesTotalEfluent"]["cabal"] = sumIgnoreNone(actualIndustry["cabal"], listOfEDARCompounds[industry[1]]["industriesTotalInfluent"]["cabal"])
                    

    wb_ptr = openpyxl.load_workbook(edar_ptr_xlsx)
    ws_ptr = wb_ptr["Hoja1"]

    for ptr in ws_ptr.iter_rows():
        # if ptr[1].value == "ES9081940001010E":
        #     print("trobat")
            if ptr[1].value in listOfEDARCompounds:
                for first_line in ws_ptr.iter_rows(max_row=1):
                    i = 0
                    for ptr_key in first_line:
                        if ptr_key.value in listOfEDARCompounds[ptr[1].value]["efluent"]:
                            if ptr[i].value != "NULL":
                               listOfEDARCompounds[ptr[1].value]["efluent"][ptr_key.value] = ptr[i].value
                        i += 1



    listEdars = {}

    # print(listOfEDARCompounds["ES9081940001010E"])

    for edar in listOfEDARCompounds:
        listEdars[listOfEDARCompounds[edar]["dc_code"]] = listOfEDARCompounds[edar]

    wb_analitiques = openpyxl.load_workbook(edar_analitiques_xlsx)
    ws_analitiques = wb_analitiques["Hoja1"]

    for an in ws_analitiques.iter_rows():

        # Totes les dades estan en mg/l
        # Cabal està en m3/dia

        if an[0].value in listEdars:
            # if an[0].value == "DBSS":
            #     print("found", an[4].value, an[6].value, an[8].value)
            if an[3].value != "NUll":
                if isANumber(an[3].value):
                    listEdars[an[0].value]["efluent"]["cabal"] = an[3].value
                else:
                    listEdars[an[0].value]["efluent"]["cabal"] = an[3].value.replace(",", ".")
            if an[5].value != "NULL":
                if isANumber(an[5].value):
                    listEdars[an[0].value]["efluent"]["dbo"] = an[5].value
                else:
                    listEdars[an[0].value]["efluent"]["dbo"] = an[5].value.replace(",", ".")
            if an[7].value != "NULL":
                if isANumber(an[7].value):
                    listEdars[an[0].value]["efluent"]["nitrogen"] = an[7].value
                else:
                    listEdars[an[0].value]["efluent"]["nitrogen"] = an[7].value.replace(",", ".")
            if an[9].value != "NULL":
                if isANumber(an[9].value):
                    listEdars[an[0].value]["efluent"]["fosfor"] = an[9].value
                else:
                    listEdars[an[0].value]["efluent"]["fosfor"] = an[9].value.replace(",", ".")
            listEdars[an[0].value]["configuration"] = []
            listEdars[an[0].value]["configuration"].append(an[10].value)
            listEdars[an[0].value]["configuration"].append(an[11].value)
            if an[12].value != "" and an[12].value is not None:
                # print(an[11].value)
                for terciari in an[12].value.split(","):
                    listEdars[an[0].value]["configuration"].append(terciari.strip())

            listEdars[an[0].value]["efluentLoad"] = {}
            for compoundEffluent in listEdars[an[0].value]["efluent"]:
                if(listEdars[an[0].value]["efluent"][compoundEffluent] != None and listEdars[an[0].value]["efluent"][compoundEffluent] != "-" and listEdars[an[0].value]["efluent"]["cabal"] != None):
                    #print(listEdars[an[0].value]["efluent"][compoundEffluent], listEdars[an[0].value]["efluent"]["cabal"])
                    listEdars[an[0].value]["efluentLoad"][compoundEffluent] = float(listEdars[an[0].value]["efluent"][compoundEffluent]) * (float(listEdars[an[0].value]["efluent"]["cabal"]) * 1000)
                else:
                    listEdars[an[0].value]["efluentLoad"][compoundEffluent] = None



    return listEdars



# data = calcAllDataForNils()
# edar = "DABR"
# print(data[edar]["efluentLoad"]["dbo"], data[edar]["industriesTotalInfluent"]["dbo"], data[edar]["efluent"]["cabal"]*1000 )
# print(data[edar]["efluentLoad"]["dbo"] / (float(data[edar]["efluent"]["cabal"])*1000))