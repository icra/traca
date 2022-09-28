
import csv
import openpyxl
import json
from sqlalchemy import create_engine


def sumIgnoreNone(object, compount, reciver):
    try:
        if compount != 'cabal':
            if object[compount] != None:
                if (isANumber(object[compount])):
                    return ((object[compount] * object["cabal"]) + (reciver[compount] * reciver["cabal"])) / (
                            object["cabal"] + reciver["cabal"])
            else:
                return ((reciver[compount] * reciver["cabal"])) / (
                        object["cabal"] + reciver["cabal"])
        else:
            if object[compount] != None:
                if (isANumber(object[compount])):
                    return reciver[compount] + object[compount]
        return reciver[compount]
    except:
        print("Error in sumIgnoreNone", compount, reciver[compount], object[compount], reciver)
        return reciver[compount]

def isANumber(myVariable):
    if type(myVariable) == int or type(myVariable) == float:
        return True
    else:
        return False

def isfloat(num):
    try:
        float(num)
        return True
    except Exception as e:
        return False

# Llegeix data edar de tots els fitxers i database i retorna {dc_code: edar}, no es fa calcul dels contaminants a efluent
def calcAllDataForNilsConcentration(industries_to_edar, contaminants_i_nutrients, edar_data_xlsx ):

    listOfEDARCompounds = {}

    wb_ptr = openpyxl.load_workbook(edar_data_xlsx)
    ws_ptr = wb_ptr["Sheet1"]
    isFirst = True

    for row in ws_ptr.iter_rows():
        # if ptr[1].value == "ES9081940001010E":
        #     print("trobat")
        if isFirst:
            isFirst = False
            continue
        listOfEDARCompounds[str(row[0].value)] = {
            "eu_code": str(row[0].value),
            "dc_code": str(row[1].value),
            "nom": str(row[2].value),
            "population_real": int(row[3].value),
            "configuration": [],
            "industriesTotalInfluent": {},
            "efluent": {},
            "influent": {},
            "cabal_observat": []
        }
        if(row[4].value is not None):
            listOfEDARCompounds[str(row[0].value)]['configuration'].append(str(row[4].value))
            if (row[5].value is not None):
                listOfEDARCompounds[str(row[0].value)]['configuration'].append(str(row[5].value))
                if (row[6].value is not None):
                    listOfEDARCompounds[str(row[0].value)]['configuration'].extend(str(row[6].value).replace(' ', '').split(','))

        edar_id = row[0].value

        if edar_id in industries_to_edar:
            group_of_dicharges = industries_to_edar[edar_id]
            discharges_to_edar = {}  # Dummy per reutilitzar funcio
            discharges_to_edar[edar_id] = []
            for discharge in group_of_dicharges.values():
                discharges_to_edar[edar_id].append(group_industries(discharge, contaminants_i_nutrients))

            grouped = suma_industries_abocament(discharges_to_edar, contaminants_i_nutrients, store_id=False)

            listOfEDARCompounds[edar_id]["industriesTotalInfluent"] = grouped[edar_id]

    return listOfEDARCompounds

#Funcio per calibrar % eliminacio i cxgx
def exportDataForNils(industries_to_edar, contaminants_i_nutrients, edar_data_xlsx, edar_analitiques_xlsx, edar_prtr_xlsx, connection, file_name = 'json_data.json'):
    listOfEDARCompounds = calcAllDataForNilsConcentration(industries_to_edar, contaminants_i_nutrients, edar_data_xlsx)

    dc_to_eu = {}

    for edar in listOfEDARCompounds:
        dc = listOfEDARCompounds[edar]["dc_code"]
        eu = listOfEDARCompounds[edar]["eu_code"]
        dc_to_eu[dc] = eu

    # llegeix observacions influent i efluent depuradora de TN, TP i DBO del fitxer edars_analitiques_sitemes_2
    wb_ptr = openpyxl.load_workbook(edar_analitiques_xlsx)
    ws_ptr = wb_ptr["Hoja1"]
    isFirst = False
    for ptr in ws_ptr.iter_rows():
        if not isFirst:
            isFirst = True
        else:
            dc_code = ptr[0].value
            if dc_code in dc_to_eu:
                eu_code = dc_to_eu[dc_code]
                if eu_code in listOfEDARCompounds:
                    for first_line in ws_ptr.iter_rows(max_row=1):
                        i = 0
                        for ptr_key in first_line:

                            if ptr_key.value == "cabal_diari":
                                if isfloat(ptr[i].value):
                                    value = float(ptr[i].value)
                                    listOfEDARCompounds[eu_code]["cabal_observat"].append(value)
                            else:

                                pollutant = ""
                                isEffluent = False
                                if "effluent" in ptr_key.value:
                                    pollutant = ptr_key.value.replace(" effluent", "")
                                    isEffluent = True
                                elif "influent" in ptr_key.value:
                                    pollutant = ptr_key.value.replace(" influent", "")

                                if pollutant in contaminants_i_nutrients:

                                    if isfloat(ptr[i].value):
                                        value = float(ptr[i].value)

                                        if isEffluent:
                                            if pollutant not in listOfEDARCompounds[eu_code]["efluent"]:
                                                listOfEDARCompounds[eu_code]["efluent"][pollutant] = []
                                            listOfEDARCompounds[eu_code]["efluent"][pollutant].append(value)

                                        else:
                                            if pollutant not in listOfEDARCompounds[eu_code]["influent"]:
                                                listOfEDARCompounds[eu_code]["influent"][pollutant] = []
                                            listOfEDARCompounds[eu_code]["influent"][pollutant].append(value)

                            i += 1

    # Llegeix fitxer prtr
    wb_ptr = openpyxl.load_workbook(edar_prtr_xlsx)
    ws_ptr = wb_ptr["Hoja1"]
    isFirst = False
    for ptr in ws_ptr.iter_rows():
        if not isFirst:
            isFirst = True
        else:
            eu_code = ptr[1].value
            if eu_code in listOfEDARCompounds:
                for first_line in ws_ptr.iter_rows(max_row=1):
                    i = 0
                    for ptr_key in first_line:
                        if ptr_key.value in contaminants_i_nutrients:

                            if isfloat(ptr[i].value):
                                value = float(ptr[i].value)

                                if ptr_key.value not in listOfEDARCompounds[eu_code]["efluent"]:
                                    listOfEDARCompounds[eu_code]["efluent"][ptr_key.value] = [value / 1000]
                                else:
                                    listOfEDARCompounds[eu_code]["efluent"][ptr_key.value].append(value / 1000)

                        i += 1

    # Edar scarce
    edar_scarce = connection.get_edar_scarce()
    for wwtp in edar_scarce:
        for contaminant in edar_scarce[wwtp]:
            if contaminant in contaminants_i_nutrients:

                values_o = edar_scarce[wwtp][contaminant]['efluent']
                if len(values_o) > 0:
                    if contaminant not in listOfEDARCompounds[wwtp]['efluent']:
                        listOfEDARCompounds[wwtp]["efluent"][contaminant] = []
                    listOfEDARCompounds[wwtp]["efluent"][contaminant].extend(values_o)

                values_i = edar_scarce[wwtp][contaminant]['influent']
                if len(values_i) > 0:
                    if contaminant not in listOfEDARCompounds[wwtp]['influent']:
                        listOfEDARCompounds[wwtp]["influent"][contaminant] = []
                    listOfEDARCompounds[wwtp]["influent"][contaminant].extend(values_i)

    for wwtp in listOfEDARCompounds:
        for compound in listOfEDARCompounds[wwtp]["efluent"]:
            listOfEDARCompounds[wwtp]["efluent"][compound] = (sum(listOfEDARCompounds[wwtp]["efluent"][compound]) / len(listOfEDARCompounds[wwtp]["efluent"][compound]))

        for compound in listOfEDARCompounds[wwtp]["influent"]:
            listOfEDARCompounds[wwtp]["influent"][compound] = (sum(listOfEDARCompounds[wwtp]["influent"][compound]) / len(listOfEDARCompounds[wwtp]["influent"][compound]))

        listOfEDARCompounds[wwtp]["cabal_observat"] = sum(listOfEDARCompounds[wwtp]["cabal_observat"]) / len(listOfEDARCompounds[wwtp]["cabal_observat"])


    with open(file_name, 'w', encoding='utf8') as outfile:
        json.dump(listOfEDARCompounds, outfile, ensure_ascii=False)

#Funcio per calibrar, llegeix fitxer review
def wwtp_info(review_xlsx, contaminants_i_nutrients, resum_eliminacio_xlsx, file_name = 'edars_pollutant_attenuation.json'):

    dict = {}

    wb_ptr = openpyxl.load_workbook(review_xlsx, data_only=True)
    ws_ptr = wb_ptr["All"]
    isFirst = False
    for ptr in ws_ptr.iter_rows():
        if not isFirst:
            isFirst = True
        else:
            for first_line in ws_ptr.iter_rows(max_row=1):
                i = 0
                for ptr_key in first_line:
                    if ptr[i].value in contaminants_i_nutrients and ptr[i].value not in dict:
                        dict[ptr[i].value] = {
                            "cxgx": [],
                            "excrecio": []
                        }


                    if ptr[i].value in contaminants_i_nutrients and ptr[i+2].value == "MEAN":
                        if isfloat(ptr[i+24].value):
                            dict[ptr[i].value]["UV"] = ptr[i+24].value
                        else:
                            dict[ptr[i].value]["UV"] = 0

                        if isfloat(ptr[i+25].value):
                            dict[ptr[i].value]["CL"] = ptr[i+25].value
                        else:
                            dict[ptr[i].value]["CL"] = 0

                        if isfloat(ptr[i + 27].value):
                            dict[ptr[i].value]["SF"] = ptr[i + 27].value
                        else:
                            dict[ptr[i].value]["SF"] = 0

                        if isfloat(ptr[i + 28].value):
                            dict[ptr[i].value]["UF"] = ptr[i + 28].value
                        else:
                            dict[ptr[i].value]["UF"] = 0

                        if isfloat(ptr[i + 29].value):
                            dict[ptr[i].value]["OTHER"] = ptr[i + 29].value
                        else:
                            dict[ptr[i].value]["OTHER"] = 0

                    if ptr[i].value in contaminants_i_nutrients and isfloat(ptr[i+7].value):
                        dict[ptr[i].value]["cxgx"].append(float(ptr[i+7].value) / 1000000)

                    if ptr[i].value in contaminants_i_nutrients and isfloat(ptr[i+8].value):
                        dict[ptr[i].value]["excrecio"].append(ptr[i+8].value)

                    if ptr[i].value in contaminants_i_nutrients and ptr[i+2].value == "PROPOSED":

                        if isfloat(ptr[i+10].value):
                            dict[ptr[i].value]["P"] = ptr[i+10].value
                        else:
                            dict[ptr[i].value]["P"] = 0

                        if isfloat(ptr[i+11].value):
                            dict[ptr[i].value]["SC"] = ptr[i + 11].value
                        else:
                            dict[ptr[i].value]["SC"] = 0

                        if isfloat(ptr[i+18].value):
                            dict[ptr[i].value]["SN"] = ptr[i + 18].value
                        else:
                            dict[ptr[i].value]["SN"] = 0

                        if isfloat(ptr[i+19].value):
                            dict[ptr[i].value]["SP"] = ptr[i + 19].value
                        else:
                            dict[ptr[i].value]["SP"] = 0

                    i += 1

    wb_ptr = openpyxl.load_workbook(resum_eliminacio_xlsx, data_only=True)
    for technology in ["SF", "O3", "GAC", "UV", "UF", "OI", "UV-H2O2"]:
        ws_ptr = wb_ptr[technology]
        isFirst = False
        for ptr in ws_ptr.iter_rows():
            if not isFirst:
                isFirst = True
            else:
                pol_name = str(ptr[0].value)

                if pol_name in contaminants_i_nutrients:
                    removal_rate = ptr[2].value

                    if pol_name not in dict:
                        dict[pol_name] = {}

                    if isfloat(removal_rate):
                        if technology == 'UV':
                            dict[pol_name]["UV"] = removal_rate

                        #elif technology == 'UV':
                        #    dict[ptr[i].value]["CL"] = removal_rate

                        elif technology == 'SF':
                            dict[pol_name]["SF"] = removal_rate

                        elif technology == 'UF':
                            dict[pol_name]["UF"] = removal_rate

    with open(file_name, 'w', encoding='utf8') as outfile:
        json.dump(dict, outfile, ensure_ascii=False)

# Afegeix la concentració de cada contaminant a l'efluent
def estimate_effluent(removal_rate, listEdars, contaminants_i_nutrients):

    # Llegim paràmetres calibrats
    calibrated_parameters = {}

    wb_ptr = openpyxl.load_workbook(removal_rate, data_only=True)
    ws_ptr = wb_ptr["Sheet1"]
    isFirst = True

    for row in ws_ptr.iter_rows():
        # if ptr[1].value == "ES9081940001010E":
        #     print("trobat")
        if isFirst:
            isFirst = False
            continue

        if row[0].value is not None:
            calibrated_parameters[str(row[0].value)] = {

            "name": str(row[0].value),
            "UV": float(row[1].value),
            "CL": float(row[2].value),
            "SF": float(row[3].value),
            "UF": float(row[4].value),
            "GAC": float(row[5].value),
            "RO": float(row[6].value),
            "AOP": float(row[7].value),
            "O3": float(row[8].value),
            "OTHER": float(row[9].value),
            "P": float(row[10].value),
            "SC": float(row[11].value),
            "SN": float(row[12].value),
            "SP": float(row[13].value),
            "generation_per_capita": float(row[14].value),
            "error_industrial": float(row[15].value),
        }

    q_per_capita = 0.242  # m3/dia/habitant
    contaminants = contaminants_i_nutrients

    for edar_key in listEdars.keys():

        try:
            wwtp = listEdars[edar_key]
            population = float(wwtp["population_real"])
            cabal_domestic = q_per_capita * population  # m3/dia

            cabal_influent_industrial = 0   #Depuradores sense industries
            if 'q' in wwtp["industriesTotalInfluent"]:
                cabal_influent_industrial = wwtp["industriesTotalInfluent"]["q"]  # m3/dia

            compounds_effluent = {
                "q": cabal_influent_industrial + cabal_domestic  # m3/dia
            }

            for contaminant in contaminants:
                load_influent_domestic = 0
                if contaminant in calibrated_parameters:
                    load_influent_domestic = population * calibrated_parameters[contaminant][
                        "generation_per_capita"] / 1000  # kg/dia

                load_influent_industrial = 0
                if contaminant in wwtp["industriesTotalInfluent"]:
                    load_influent_industrial = calibrated_parameters[contaminant]["error_industrial"] * wwtp["industriesTotalInfluent"][
                                                   contaminant] / 1000  # kg/dia

                load_influent_filtered = load_influent_industrial + load_influent_domestic

                if contaminant in calibrated_parameters:    #Si no tenim dades de eliminacio, assumim que no neteja res
                    for configuration in wwtp["configuration"]:
                        load_influent_filtered *= (1 - (calibrated_parameters[contaminant][configuration] / 100))

                        #load_influent_domestic *= (1 - (calibrated_parameters[contaminant][configuration] / 100))
                        #load_influent_industrial *= (1 - (calibrated_parameters[contaminant][configuration] / 100))


                compounds_effluent[contaminant] = load_influent_filtered  # kg

                """
                compounds_effluent[contaminant] = {
                    "domestic": load_influent_domestic,
                    "industrial": load_influent_industrial
                }
                """

            #Fins ara teniem TN i TP agrupant segons separacio estandard, ara separem segons tractament depuradora (mes precis)


            if "Nitrogen Total" in contaminants_i_nutrients:
                tn = compounds_effluent["Nitrogen Total"]
                if 'SC' in wwtp["configuration"]:
                    compounds_effluent["Nitrogen orgànic"] = tn * 0.03
                    compounds_effluent["Nitrats"] = tn * 0
                    compounds_effluent["Amoni"] = tn * 0.97
                elif 'SN' in wwtp["configuration"] or 'SP' in wwtp["configuration"]:
                    compounds_effluent["Nitrogen orgànic"] = tn * 0.03
                    compounds_effluent["Nitrats"] = tn * 0.66
                    compounds_effluent["Amoni"] = tn * 0.31
            if "Fòsfor total" in contaminants_i_nutrients:
                tp = compounds_effluent["Fòsfor total"]
                compounds_effluent["Fòsfor orgànic"] = tp * 0.07
                compounds_effluent["Fosfats"] = tp * 0.93


            """
            for dom_ind in ["domestic", "industrial"]:
                try:
                    if "Nitrogen Total" in contaminants_i_nutrients:
                        tn = compounds_effluent["Nitrogen Total"][dom_ind]
                        if 'SC' in wwtp["configuration"]:
                            compounds_effluent["Nitrogen orgànic"][dom_ind] = tn * 0.03
                            compounds_effluent["Nitrats"][dom_ind] = tn * 0
                            compounds_effluent["Amoni"][dom_ind] = tn * 0.97
                        elif 'SN' in wwtp["configuration"] or 'SP' in wwtp["configuration"]:
                            compounds_effluent["Nitrogen orgànic"][dom_ind] = tn * 0.03
                            compounds_effluent["Nitrats"][dom_ind] = tn * 0.66
                            compounds_effluent["Amoni"][dom_ind] = tn * 0.31
                except Exception as e:
                    pass
                try:
                    if "Fòsfor total" in contaminants_i_nutrients:
                        tp = compounds_effluent["Fòsfor total"][dom_ind]
                        compounds_effluent["Fòsfor orgànic"][dom_ind] = tp * 0.07
                        compounds_effluent["Fosfats"][dom_ind] = tp * 0.93
                except Exception as e:
                    pass
            """
            listEdars[edar_key]['compounds_effluent'] = compounds_effluent

        except Exception as e:
            print(edar_key + " missing field: " + str(e))

    return listEdars

def read_edars(contaminants_i_nutrients, industries_to_edar, edar_data_xlsx, removal_rate, swat_to_edar_code):

    edars_calibrated = calcAllDataForNilsConcentration(industries_to_edar, contaminants_i_nutrients, edar_data_xlsx)

    edars_calibrated = estimate_effluent(removal_rate, edars_calibrated, contaminants_i_nutrients)

    wb_ptr = openpyxl.load_workbook(swat_to_edar_code)
    ws_ptr = wb_ptr["Sheet1"]
    isFirst = True

    for row in ws_ptr.iter_rows():
        # if ptr[1].value == "ES9081940001010E":
        #     print("trobat")
        if isFirst:
            isFirst = False
            continue
        elif row[6].value is not None:
            edars_calibrated[row[6].value]['id_swat'] = row[0].value
            edars_calibrated[row[6].value]['lat'] = float(row[3].value)
            edars_calibrated[row[6].value]['long'] = float(row[4].value)

    return edars_calibrated

def readListOfIndustriesFromCSV(industrial_data):
    # Reads the first column of the csv file with the industrial to river mapping
    # Returns a list of strings
    industries = {}

    wb_ptr = openpyxl.load_workbook(industrial_data)
    ws_ptr = wb_ptr["Sheet1"]
    isFirst = True

    for row in ws_ptr.iter_rows():
        # if ptr[1].value == "ES9081940001010E":
        #     print("trobat")
        if isFirst:
            isFirst = False
            continue

        if row[0].value+' '+row[1].value not in industries:
            if row[2].value is not None:
                industries[row[0].value+' '+row[1].value] = {
                    "name": row[0].value+' '+row[1].value,
                    "activitat": row[0].value,
                    "abocament": row[1].value,
                    "point": int(row[2].value)
                }
    return industries

def group_industries(abocaments_activitat_en_una_ubicacio, contaminants_i_nutrients):

    aux_point = {
        "activitat": abocaments_activitat_en_una_ubicacio[0]["activitat/ubicacio"],
        "abocament": abocaments_activitat_en_una_ubicacio[0]["nom_abocament"],
        "q": 0,
    }

    for compound in abocaments_activitat_en_una_ubicacio:

        if compound["valor_minim"] is not None and compound["valor_maxim"] is not None:
            valor_mitja = (float(compound["valor_minim"]) + float(compound["valor_maxim"])) / 2
        elif compound["valor_maxim"] is not None:
            valor_mitja = float(compound["valor_maxim"])
        elif compound["valor_minim"] is not None:
            valor_mitja = float(compound["valor_minim"])

        """
        Passem tot a mg/l (tots els contaminents que tractem estan en µg/l, ng/l o mg/l, el cabal 
        esta en m3/dia o m3/any i no es veu afectat)
        """
        unitats = compound["unitats"]
        if unitats is not None:
            if "µg/l" in unitats:
                valor_mitja = valor_mitja / 1000
            elif "ng/l" in unitats:
                valor_mitja = valor_mitja / 1000000


        if compound["nom_variable"] == "Cabal diari":
            if aux_point["q"] == 0 and compound["valor_maxim"] is not None:
                aux_point["q"] = valor_mitja
        elif compound["nom_variable"] == "Cabal anual" and compound["valor_maxim"] is not None:
            aux_point["q"] = valor_mitja / 365
        elif compound["nom_variable"] in contaminants_i_nutrients:
            aux_point[compound["nom_variable"]] = valor_mitja

    #Passar a carrega
    for key in aux_point:
        if key not in ['activitat', 'abocament', 'q']:
            aux_point[key] = float(aux_point[key]) * float(aux_point["q"])


    if "Nitrats" not in aux_point:
        if "Nitrogen Total" in aux_point and "Amoni" in aux_point and "Nitrogen orgànic" in aux_point:
            nitrats = aux_point["Nitrogen Total"] - aux_point["Amoni"] - aux_point["Nitrogen orgànic"]
            if nitrats < 0:
                nitrats = 0
            aux_point["Nitrats"] = nitrats

        elif "Nitrogen Total" in aux_point:
            aux_point["Nitrats"] = aux_point["Nitrogen Total"]*0.1

    if "Amoni" not in aux_point:
        if "Nitrogen Total" in aux_point and "Nitrats" in aux_point and "Nitrogen orgànic" in aux_point:
            amoni = aux_point["Nitrogen Total"] - aux_point["Nitrats"] - aux_point["Nitrogen orgànic"]
            if amoni < 0:
                amoni = 0
            aux_point["Amoni"] = amoni
        elif "Nitrogen Kjeldahl" in aux_point and "Nitrogen orgànic" in aux_point:
            amoni = aux_point["Nitrogen Kjeldahl"] - aux_point["Nitrogen orgànic"]
            if amoni < 0:
                amoni = 0
            aux_point["Amoni"] = amoni

        elif "Nitrogen Total" in aux_point:
            aux_point["Amoni"] = aux_point["Nitrogen Total"]*0.87
        elif "Nitrogen Kjeldahl" in aux_point:
            aux_point["Amoni"] = aux_point["Nitrogen Kjeldahl"]*0.9
    if "Nitrogen orgànic" not in aux_point:
        if "Nitrogen Total" in aux_point and "Nitrats" in aux_point and "Amoni" in aux_point:
            n_org = aux_point["Nitrogen Total"] - aux_point["Nitrats"] - aux_point["Amoni"]
            if n_org < 0:
                n_org = 0
            aux_point["Nitrogen orgànic"] = n_org
        elif "Nitrogen Kjeldahl" in aux_point and "Amoni" in aux_point:
            n_org = aux_point["Nitrogen Kjeldahl"] - aux_point["Amoni"]
            if n_org < 0:
                n_org = 0
            aux_point["Nitrogen orgànic"] = n_org

        elif "Nitrogen Total" in aux_point:
            aux_point["Nitrogen orgànic"] = aux_point["Nitrogen Total"]*0.03
        elif "Nitrogen Kjeldahl" in aux_point:
            aux_point["Nitrogen orgànic"] = aux_point["Nitrogen Kjeldahl"]*0.1
    if "Fosfats" not in aux_point:
        if "Fòsfor Total" in aux_point and "Fòsfor orgànic" in aux_point:

            po3 = aux_point["Fòsfor Total"] - aux_point["Fòsfor orgànic"]
            if po3 < 0:
                po3 = 0
            aux_point["Fosfats"] = po3

        elif "Fòsfor Total" in aux_point:
            aux_point["Fosfats"] = aux_point["Fòsfor Total"]*0.93
    if "Fòsfor orgànic" not in aux_point:
        if "Fòsfor Total" in aux_point and "Fosfats" in aux_point:

            p_org = aux_point["Fòsfor Total"] - aux_point["Fosfats"]
            if p_org < 0:
                p_org = 0
            aux_point["Fòsfor orgànic"] = p_org

        elif "Fòsfor Total" in aux_point:
            aux_point["Fòsfor orgànic"] = aux_point["Fòsfor Total"]*0.07
    if "Nitrogen Total" not in aux_point:
        nitrogen_total = 0
        if "Nitrats" in aux_point:
            nitrogen_total += aux_point["Nitrats"]
        if "Nitrogen orgànic" in aux_point:
            nitrogen_total += aux_point["Nitrogen orgànic"]
        if "Amoni" in aux_point:
            nitrogen_total += aux_point["Amoni"]
        aux_point["Nitrogen Total"] = nitrogen_total
    if "Fòsfor Total" not in aux_point:
        fosfor_total = 0
        if "Fosfats" in aux_point:
            fosfor_total += aux_point["Fosfats"]
        if "Fòsfor orgànic" in aux_point:
            fosfor_total += aux_point["Fòsfor orgànic"]
        aux_point["Fòsfor Total"] = fosfor_total

    return aux_point

def nom_abocament_a_id(industrial_data_file, recall_points_file):

    #Llegir industrial data
    industries = {}
    wb_ptr = openpyxl.load_workbook(industrial_data_file)
    ws_ptr = wb_ptr["Sheet1"]
    isFirst = True
    for row in ws_ptr.iter_rows():
        if isFirst:
            isFirst = False
            continue
        if row[2].value is not None:
            industries[row[1].value] = int(row[2].value)

    #Llegir recall data
    point_2_id = {}
    wb_ptr = openpyxl.load_workbook(recall_points_file)
    ws_ptr = wb_ptr["Sheet1"]
    isFirst = True
    for row in ws_ptr.iter_rows():
        if isFirst:
            isFirst = False
            continue

        if row[7].value is not None:
            if row[7].value not in point_2_id:
                point_2_id[int(row[7].value)] = row[0].value

    discharge_point_to_id = {}
    for discharge_point, ind_id in industries.items():
        if ind_id in point_2_id:
            discharge_point_to_id[discharge_point] = point_2_id[ind_id]
    return discharge_point_to_id

def suma_industries_abocament(abocaments, contaminants_i_nutrients, store_id = True):
    compounds = contaminants_i_nutrients.copy()
    compounds.append("q")   #Sumem càrregues d'abocaments + cabal
    abocaments_sumat = {}
    for id_abocament, abocament in abocaments.items():
        if store_id:
            aux = {
                "id": id_abocament,
                "abocament": abocament[0]["abocament"]
            }
        else:
            aux = {}

        for industry in abocament:
            for compound in industry:
                if compound in compounds:
                    if compound not in aux:
                        aux[compound] = 0
                    aux[compound] += industry[compound]

        abocaments_sumat[id_abocament] = aux
    return abocaments_sumat

def read_industries(industries_to_river, industrial_data_file, recall_points_file, contaminants_i_nutrients, connection):

    industries_grouped = {}

    # industries_to_river[activitat + abocament]: {dades abocament complert}
    for key, industry in industries_to_river.items():
        industries_grouped[key] = group_industries(industry, contaminants_i_nutrients)


    #Llegir dades de industrial_data i recall_points
    discharge_point_to_id = nom_abocament_a_id(industrial_data_file, recall_points_file)

    #Agrupar les industries en punts d'abocament
    discharge_points = {}
    for industry in industries_grouped.values():
        nom_abocament = industry['abocament']
        if nom_abocament in discharge_point_to_id:
            id_abocament = discharge_point_to_id[nom_abocament]
            if not id_abocament in discharge_points:
                discharge_points[id_abocament] = []
            discharge_points[id_abocament].append(industry)

    contaminants_per_punt_abocament = suma_industries_abocament(discharge_points, contaminants_i_nutrients)

    #concentracions_avg = {}
    #for contaminant in contaminants_i_nutrients:
    #    concentracions_avg[contaminant] = connection.avg_estacions_riu(contaminant)

    for contaminant in contaminants_i_nutrients:
        for abocament in contaminants_per_punt_abocament:
            #contaminants_per_punt_abocament[abocament][contaminant] = contaminants_per_punt_abocament[abocament]["q"] * concentracions_avg[contaminant] * 10 / 1000  # Passem a kg

            if contaminant in contaminants_per_punt_abocament[abocament]:
                contaminants_per_punt_abocament[abocament][contaminant] = contaminants_per_punt_abocament[abocament][contaminant] / 1000 #Passem a kg


    return contaminants_per_punt_abocament