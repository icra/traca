from lib.db.Custom_SQLite import Custom_SQLite as cS
from lib.db.ConnectPostgree import ConnectDb as pg

import csv
import openpyxl


def sumIgnoreNone(object, compount, reciver):
    try:
        if compount != 'cabal':
            if object[compount] != None:
                if (isANumber(object[compount])):
                    return ((object[compount] * object["cabal"]) + (reciver[compount] * reciver["cabal"])) / (
                            object["cabal"] + reciver["cabal"])
            else:
                return ((reciver[compount] * reciver["cabal"])) / (
                        object["cabal"] + reciver["cabal"])
        else:
            if object[compount] != None:
                if (isANumber(object[compount])):
                    return reciver[compount] + object[compount]
        return reciver[compount]
    except:
        print("Error in sumIgnoreNone", compount, reciver[compount], object[compount], reciver)
        return reciver[compount]


def isANumber(myVariable):
    if type(myVariable) == int or type(myVariable) == float:
        return True
    else:
        return False

# Llegeix data edar de tots els fitxers i database i retorna {dc_code: edar}, no es fa calcul dels contaminants a efluent
def calcAllDataForNilsConcentration(edar_compounds_csv, edar_population_csv, edar_analitiques_xlsx, edar_ptr_xlsx, edar_cabals_xlsx):
    pg_url = "icra.loading.net"
    pg_user = "traca_user"
    pg_pass = "EdificiH2O!"
    pg_db = "traca_1"

    connection = pg(pg_url, pg_db, pg_user, pg_pass)
    industries = connection.getIndustriesToEdar()

    # Fields of edar compounts
    # cabal_diari m3/dia
    # unidades mg/l
    listOfEDARCompounds = {}
    with open(edar_compounds_csv, encoding='utf8', newline='') as csvEDARCompounts:
        reader = csv.reader(csvEDARCompounts, delimiter=';')
        isFirst = True
        for row in reader:
            if isFirst:
                isFirst = False
            else:
                listOfEDARCompounds[row[0]] = {
                    "eu_code": row[0],
                    "dc_code": row[1],
                    "nom": row[2],
                    "industries": [],
                    "industriesTotalInfluent": {
                        "amoni": 0,
                        "dbo": 0,
                        "fosfats": 0,
                        "toc": 0,
                        "fosfor": 0,
                        "nitrats": 0,
                        "nitrogen_org": 0,
                        "cabal": 0,
                    },
                    "industriesTotalEffluent": {
                        "amoni": 0,
                        "dbo": 0,
                        "fosfats": 0,
                        "toc": 0,
                        "fosfor": 0,
                        "nitrats": 0,
                        "nitrogen_org": 0,
                        "cabal": 0,
                    },
                    "configuration": [],
                    "efluent": {
                        "dbo": 0,
                        "nitrogen": 0,
                        "fosfor": 0,
                        "arsènic i compostos com as µg as/l": None,
                        "cadmi i compostos com cd µg cd/l": None,
                        "crom i compostos com cr µg cr/l": None,
                        "coure i compostos com cu µg cu/l": None,
                        "mercuri i compostos com hg µg hg/l": None,
                        "niquel i compostos com ni µg ni/l": None,
                        "plom i compostos com pb µg pb/l": None,
                        "zenc i compostos com zn µg zn/l": None,
                        "cianurs com cn total µg cn/l": None,
                        "fluorurs com f total µg f/l": None,
                        "clorurs com cl total µg cl/l": None,
                        "1,2 dicloroetà µg/l": None,
                        "diclormetà µg/l": None,
                        "tetracloretilè per µg/l": None,
                        "tetraclormetà tcm tetracloruro de carbono µg/l": None,
                        "tricloretilètri µg/l": None,
                        "triclormetà cloroformo µg/l": None,
                        "hexaclorbenzè µg/l": None,
                        "lindà gamma-hch µg/l": None,
                        "pentaclorfenol pcp µg/l": None,
                        "policlorbifenils pcbs µg/l": None,
                        "benzè µg/l": None,
                        "toluè µg/l": None,
                        "xilens µg/l": None,
                        "naftalè µg/l": None,
                        "fluorantè µg/l": None,
                        "benzo g, h, i, perilè µg/l": None,
                        "suma hap µg/l": None,
                        "atrazina µg/l": None,
                        "simazina µg/l": None,
                        "isoproturon µg/l": None,
                        "diuron µg/l": None,
                        "nonilfenol etoxilats np/np1,2eo µg/l": None,
                        "octilfenol etoxilatsop/op1,2eo µg/l": None,
                        "ftalat de bis 2-etilhexil deph µg/l": None,
                        "index de fenols µg fenol/l": None,
                        "compostos orgànics halogenats aox µgcl / l": None,
                        "tributil estany tbt µg tbt/l": None,
                        "tributil estany i els seus compostos mbt+dbt+tbt com tbt µg tb": None,
                        "trifenil estany tpht µg tpht/l": None,
                        "trifenil estany i els seus compostosmpht+dpht+tpht comtpht µg": None,
                        "compostos organoestannics totalscom sn µg sn/l": None,
                        "conductivitat 2Noneºc µs/cm": None,
                        "m+p-xilè µg/l": None,
                        "o-xilè µg/l": None,
                        "benzoapirè µg/l": None,
                        "benzobfluorantè µg/l": None,
                        "benzokfluorantè µg/l": None,
                        "indè1.2.3pirè µg/l": None,
                        "pcb 28 µg/l": None,
                        "pcb 52 µg/l": None,
                        "pcb 1None1 µg/l": None,
                        "pcb 118 µg/l": None,
                        "pcb 153 µg/l": None,
                        "pcb 138 µg/l": None,
                        "pcb 18None µg/l": None,
                        "ph u. ph": None,
                        "toc µg/l": None,
                        "dqo no decantada µg/l": None,
                        "nitrogen kjeldahl µg/l n": None,
                        "nitrogen total µg/l n": None,
                        "nitrats + nitrits µg/l n": None,
                        "4-nonilfenol dietoxilado": None,
                        "4-nonilfenol monoetoxilado": None,
                        "4-octilfenol dietoxilado": None,
                        "4-octilfenol monoetoxilado µg/l": None,
                        "compuestos organoestánnicos µg/l": None,
                        "dibutilestaño µg/l": None,
                        "difenilestaño µg/l": None,
                        "monobutilestaño µg/l": None,
                        "monofenilestaño µg/l": None,
                        "trifenilestaño µg/l": None,
                        "npeoe=None": None,
                        "t-octilfenol": None,
                        "dbo 5 dies µg o2/l": None,
                        "propazina µg/l": None,
                        "tertbutilazina µg/l": None,
                        "terbutrin µg/l": None,
                        "cibutrina µg/l": None,
                        "suma triazines µg/l": None,
                        "fòsfor total µg p/l": None
                    },
                    "influent": {
                        "dbo": None,
                        "nitrogen": None,
                        "fosfor": None
                    }
                }

    #

    with open(edar_population_csv, encoding='utf8', newline='') as csvEDARPopulation:
        reader = csv.reader(csvEDARPopulation, delimiter=',')
        isFirst = True
        for row in reader:
            if isFirst:
                isFirst = False
            else:
                if row[0] in listOfEDARCompounds:
                    listOfEDARCompounds[row[0]]["population"] = row[6]
                    listOfEDARCompounds[row[0]]["pe"] = row[1]

    # ('FOMENTO AGRICOLA Y GANADERO, SL', 'ES9080010001010E', None, 750.0, None, None, 50.0, None, None, 651.0)

    for industry in industries:
        if industry[1] in listOfEDARCompounds:
            if industry[10] is not None and industry[10] > 0:

                actualIndustry = {
                    "nom": industry[0],
                    "tipus": industry[2],
                    "amoni": industry[3],
                    "dbo": industry[4],
                    "fosfats": industry[5],
                    "toc": industry[6],
                    "fosfor": industry[7],
                    "nitrats": industry[8],
                    "nitrogen_org": industry[9],
                    "cabal": industry[10],
                }

                # print(listOfEDARCompounds[industry[1]]["industries"])

                if actualIndustry["amoni"] is not None:
                    actualIndustry["amoni"] = industry[3]
                if actualIndustry["dbo"] is not None:
                    actualIndustry["dbo"] = industry[4]
                if actualIndustry["fosfats"] is not None:
                    actualIndustry["fosfats"] = industry[5]
                if actualIndustry["toc"] is not None:
                    actualIndustry["toc"] = industry[6]
                if actualIndustry["fosfor"] is not None:
                    actualIndustry["fosfor"] = industry[7]
                if actualIndustry["nitrats"] is not None:
                    actualIndustry["nitrats"] = industry[8]
                if actualIndustry["nitrogen_org"] is not None:
                    actualIndustry["nitrogen_org"] = industry[9]
                if actualIndustry["cabal"] is not None:
                    actualIndustry["cabal"] = industry[10]

                listOfEDARCompounds[industry[1]]["industries"].append(actualIndustry)

                if industry[2] in ["Abocament", "Depuradora", "Entrada EDAR"]:
                    for compound in listOfEDARCompounds[industry[1]]["industriesTotalInfluent"]:
                        # print(industry[1], listOfEDARCompounds[industry[1]]["industriesTotalInfluent"])
                        iti = listOfEDARCompounds[industry[1]]["industriesTotalInfluent"]
                        iti[compound] = sumIgnoreNone(actualIndustry, compound, iti)
                        # if industry[1] == "ES9080010001010E" and compound in ["dbo", "cabal"]:
                        #     print(actualIndustry)
                        #     print("-- "+str(iti["dbo"]) +" "+str(iti["cabal"]))
                else:
                    for compound in listOfEDARCompounds[industry[1]]["industriesTotalEffluent"]:
                        iti = listOfEDARCompounds[industry[1]]["industriesTotalEffluent"]
                        iti[compound] = sumIgnoreNone(actualIndustry, compound, iti)

    wb_ptr = openpyxl.load_workbook(edar_ptr_xlsx)
    ws_ptr = wb_ptr["Hoja1"]

    for ptr in ws_ptr.iter_rows():
        # if ptr[1].value == "ES9081940001010E":
        #     print("trobat")
        if ptr[1].value in listOfEDARCompounds:
            for first_line in ws_ptr.iter_rows(max_row=1):
                i = 0
                for ptr_key in first_line:
                    if ptr_key.value in listOfEDARCompounds[ptr[1].value]["efluent"]:
                        if ptr[i].value != "NULL":
                            listOfEDARCompounds[ptr[1].value]["efluent"][ptr_key.value] = ptr[i].value
                    i += 1

    listEdars = {}

    # print(listOfEDARCompounds["ES9081940001010E"])

    for edar in listOfEDARCompounds:
        listEdars[listOfEDARCompounds[edar]["dc_code"]] = listOfEDARCompounds[edar]

    wb_analitiques = openpyxl.load_workbook(edar_analitiques_xlsx)
    ws_analitiques = wb_analitiques["Hoja1"]

    for an in ws_analitiques.iter_rows():

        # Totes les dades estan en mg/l
        # Cabal està en m3/dia

        if an[0].value in listEdars:
            # if an[0].value == "DBSS":
            #     print("found", an[4].value, an[6].value, an[8].value)

            if an[3].value != "NUll":
                if isANumber(an[3].value):
                    listEdars[an[0].value]["efluent"]["cabal"] = an[3].value
                else:
                    listEdars[an[0].value]["efluent"]["cabal"] = an[3].value.replace(",", ".")

            if an[4].value != "NULL":
                if isANumber(an[4].value):
                    listEdars[an[0].value]["influent"]["dbo"] = an[4].value
                else:
                    listEdars[an[0].value]["influent"]["dbo"] = an[4].value.replace(",", ".")
            if an[5].value != "NULL":
                if isANumber(an[5].value):
                    listEdars[an[0].value]["efluent"]["dbo"] = an[5].value
                else:
                    listEdars[an[0].value]["efluent"]["dbo"] = an[5].value.replace(",", ".")
            if an[6].value != "NULL":
                if isANumber(an[6].value):
                    listEdars[an[0].value]["influent"]["nitrogen"] = an[6].value
                else:
                    listEdars[an[0].value]["influent"]["nitrogen"] = an[6].value.replace(",", ".")
            if an[7].value != "NULL":
                if isANumber(an[7].value):
                    listEdars[an[0].value]["efluent"]["nitrogen"] = an[7].value
                else:
                    listEdars[an[0].value]["efluent"]["nitrogen"] = an[7].value.replace(",", ".")
            if an[8].value != "NULL":
                if isANumber(an[8].value):
                    listEdars[an[0].value]["influent"]["fosfor"] = an[8].value
                else:
                    listEdars[an[0].value]["influent"]["fosfor"] = an[8].value.replace(",", ".")
            if an[9].value != "NULL":
                if isANumber(an[9].value):
                    listEdars[an[0].value]["efluent"]["fosfor"] = an[9].value
                else:
                    listEdars[an[0].value]["efluent"]["fosfor"] = an[9].value.replace(",", ".")
            listEdars[an[0].value]["configuration"] = []
            listEdars[an[0].value]["configuration"].append(an[10].value)
            listEdars[an[0].value]["configuration"].append(an[11].value)
            if an[12].value != "" and an[12].value is not None:
                # print(an[11].value)
                for terciari in an[12].value.split(","):
                    listEdars[an[0].value]["configuration"].append(terciari.strip())

            listEdars[an[0].value]["efluentLoad"] = {}
            for compoundEffluent in listEdars[an[0].value]["efluent"]:
                if (listEdars[an[0].value]["efluent"][compoundEffluent] != None and listEdars[an[0].value]["efluent"][
                    compoundEffluent] != "-" and listEdars[an[0].value]["efluent"]["cabal"] != None):
                    # print(listEdars[an[0].value]["efluent"][compoundEffluent], listEdars[an[0].value]["efluent"]["cabal"])
                    listEdars[an[0].value]["efluentLoad"][compoundEffluent] = float(
                        listEdars[an[0].value]["efluent"][compoundEffluent]) * (float(
                        listEdars[an[0].value]["efluent"]["cabal"]) * 1000)
                else:
                    listEdars[an[0].value]["efluentLoad"][compoundEffluent] = None

    wb_populations = openpyxl.load_workbook(edar_cabals_xlsx)
    ws_populations = wb_populations["Full2"]
    for population in ws_populations.iter_rows():
        if population[6].value in listEdars:
            # print(population[6].value, population[12].value)
            if population[12].value != "" and population[12].value is not None:
                listEdars[population[6].value]["population_real"] = population[12].value
                listEdars[population[6].value]["q_estimated"] = population[12].value * 0.242
                listEdars[population[6].value]["dbo_load_estimated_influent"] = (population[12].value * 60) + \
                                                                                (listEdars[population[6].value][
                                                                                     "industriesTotalInfluent"]["dbo"] *
                                                                                 listEdars[population[6].value][
                                                                                     "industriesTotalInfluent"][
                                                                                     "cabal"])
                if listEdars[population[6].value]["efluent"]["cabal"] is not None and \
                        listEdars[population[6].value]["efluent"]["cabal"] != "-" and \
                        listEdars[population[6].value]["efluent"]["cabal"] != "":
                    listEdars[population[6].value]["influent"]["dbo_load"] = float(
                        listEdars[population[6].value]["efluent"]["cabal"]) * float(
                        listEdars[population[6].value]["influent"]["dbo"])
            else:
                listEdars[population[6].value]["population_real"] = 1
                listEdars[population[6].value]["q_estimated"] = 1 * 0.242

    return listEdars

# retorna diccionari amb edars {eu_code: edar}, no es fa calcul dels contaminants a efluent
def edarsCalibratedFormated(edar_compounds_csv, edar_population_csv, edar_analitiques_xlsx, edar_ptr_xlsx, edar_cabals_xlsx):
    edars = calcAllDataForNilsConcentration(edar_compounds_csv, edar_population_csv, edar_analitiques_xlsx, edar_ptr_xlsx, edar_cabals_xlsx)
    edars_formatted = {}
    for edar in edars.values():
        edars_formatted[edar['eu_code']] = edar
    return edars_formatted

# Afegeix la concentració de cada contaminant a l'efluent
def estimate_effluent(removal_rate_csv, listEdars):

    # Llegim paràmetres calibrats
    calibrated_parameters = {}
    with open(removal_rate_csv, encoding='utf8', newline='') as csvfile:
        isFirst = True
        reader = csv.reader(csvfile, delimiter=';')
        for row in reader:
            if isFirst:
                isFirst = False
            else:
                compound_id, name, generation_per_capita, primary, sp, sn, sc, uf, cl, uv, other, sf = row
                calibrated_parameters[compound_id] = {
                    "name": name,
                    "generation_per_capita": float(generation_per_capita.replace(",", ".")),
                    "P": float(primary.replace(",", ".")),
                    "SP": float(sp.replace(",", ".")),
                    "SN": float(sn.replace(",", ".")),
                    "SC": float(sc.replace(",", ".")),
                    "UF": float(uf.replace(",", ".")),
                    "CL": float(cl.replace(",", ".")),
                    "UV": float(uv.replace(",", ".")),
                    "OTHER": float(other.replace(",", ".")),
                    "SF": float(sf.replace(",", ".")),
                }


    q_per_capita = 0.242  # m3/dia/habitant
    contaminants = ["dbo", "fosfor", "nitrogen"]

    for edar_key in listEdars.keys():

        try:
            wwtp = listEdars[edar_key]
            population = float(wwtp["population_real"])
            cabal_domestic = q_per_capita * population  # m3/dia
            cabal_influent_industrial = wwtp["industriesTotalInfluent"]["cabal"]  # m3/dia
            cabal_efluent_industrial = wwtp["industriesTotalEffluent"]["cabal"]  # m3/dia
            compounds_effluent = {
                "cabal": wwtp["industriesTotalInfluent"]["cabal"] + wwtp["industriesTotalEffluent"][
                    "cabal"] + cabal_domestic  # m3/dia
            }

            for contaminant in contaminants:
                if contaminant == "nitrogen":  # nitrogen
                    load_influent_domestic = population * calibrated_parameters[contaminant][
                        "generation_per_capita"] / 1000  # kg/dia

                    load_influent_industrial = 0  # De moment no agafem dades d'industria
                    load_influent_filtered = load_influent_industrial + load_influent_domestic
                    for configuration in wwtp["configuration"]:
                        load_influent_filtered *= (1 - (calibrated_parameters[contaminant][configuration] / 100))

                    n_to_components = {
                        "nitrogen_org": 1/3,
                        "amoni": 1/3,
                        "nitrats": 1/3,
                    }
                    if "SN" in wwtp["configuration"] or "SP" in wwtp["configuration"]:
                        #if(((load_influent_filtered * 1000)  / cabal_domestic) > 15): print("concentracio TN ilegal")
                        n_to_components["nitrogen_org"] = 0.03
                        n_to_components["nitrats"] = 0.66
                        n_to_components["amoni"] = 0.31
                    elif "SC" in wwtp["configuration"]:
                        n_to_components["nitrogen_org"] = 0.03
                        n_to_components["nitrats"] = 0
                        n_to_components["amoni"] = 0.97

                    # EL TN que surt de la depuradora el separem per components
                    load_influent_filtered_amoni = load_influent_filtered * n_to_components["amoni"]
                    load_influent_filtered_nitrats = load_influent_filtered * n_to_components["nitrats"]
                    load_influent_filtered_nitrogen_org = load_influent_filtered * n_to_components["nitrogen_org"]

                    """
                    load_efluent_industrial_amoni = wwtp["industriesTotalEffluent"][
                                                        "amoni"] * cabal_efluent_industrial / 1000  # kg/dia
                    load_efluent_industrial_nitrats = wwtp["industriesTotalEffluent"][
                                                          "nitrats"] * cabal_efluent_industrial / 1000  # kg/dia
                    load_efluent_industrial_nitrogen_org = wwtp["industriesTotalEffluent"][
                                                               "nitrogen_org"] * cabal_efluent_industrial / 1000  # kg/dia
                    """
                    compounds_effluent["amoni"] = load_influent_filtered_amoni  # kg
                    compounds_effluent["nitrats"] = load_influent_filtered_nitrats  # kg
                    compounds_effluent["nitrogen_org"] = load_influent_filtered_nitrogen_org  # kg

                else:  # resta de contaminants
                    load_influent_domestic = population * calibrated_parameters[contaminant][
                        "generation_per_capita"] / 1000  # kg/dia
                    load_influent_industrial = wwtp["industriesTotalInfluent"][
                                                   contaminant] * cabal_influent_industrial / 1000  # kg/dia
                    load_efluent_industrial = wwtp["industriesTotalEffluent"][
                                                  contaminant] * cabal_efluent_industrial / 1000  # kg/dia
                    load_influent_filtered = load_influent_industrial + load_influent_domestic
                    for configuration in wwtp["configuration"]:
                        load_influent_filtered *= (1 - (calibrated_parameters[contaminant][configuration] / 100))

                    compounds_effluent[contaminant] = (load_influent_filtered + load_efluent_industrial)  # kg

            listEdars[edar_key]['compounds_effluent'] = compounds_effluent

        except Exception as e:
            print(edar_key + " missing field: " + str(e))

    return listEdars

# Retorna {edar_code: edar}, on edar ja te tota la informacio
def read_edars(swat_to_edar_code_csv, edar_compounds_csv, edar_population_csv, edar_analitiques_xlsx, edar_ptr_xlsx, edar_cabals_xlsx, removal_rate_csv):

    edars_calibrated = edarsCalibratedFormated(edar_compounds_csv, edar_population_csv, edar_analitiques_xlsx, edar_ptr_xlsx, edar_cabals_xlsx)

    edars_calibrated = estimate_effluent(removal_rate_csv, edars_calibrated)

    with open(swat_to_edar_code_csv) as csv_file:
        csv_reader = csv.reader(csv_file, delimiter=',')
        line_count = 0
        for row in csv_reader:
            if line_count == 0:
                line_count += 1
            else:
                if row[8] != '':
                    edars_calibrated[row[8]]['nom_swat'] = row[1]
                    edars_calibrated[row[8]]['lat'] = row[3]
                    edars_calibrated[row[8]]['long'] = row[4]

    return edars_calibrated
